import re
from time import sleep
from urllib.parse import urlencode

import openpyxl
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import unittest


class information_gather():

    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=r'F:\software\python\python3.7\Tools\geckodriver\chromedriver.exe')
        # self.driver = webdriver.Firefox(
        #     executable_path=r'F:\software\python\python3.7\Tools\geckodriver\geckodriver.exe')

        self.information_list = []
        # self.invested_enterprise_information_list = [['308064a33078fcff29dfd220d4e3dd85', '北京三快科技有限公司'],
        #                                              ['h4e7dd628b1d115d6d9af63b5c3c50a8', 'Ant Troop HK Limited'],
        #                                              ['h601548f070c5fe41b19291b45a694d2', '美團財富有限公司'],
        #                                              ['h74f97246bac17897c5970e69515b853', 'Solid Bit Hong Kong Limited'],
        #                                              ['hf99e8059779c17945a857d068517bd8', '小象零售有限公司'],
        #                                              ['1586876a15fbbc5046f5ab0f09a4d262', '北京三快在线科技有限公司'],
        #                                              ['h76bf1f4fff88749b121c0bab1c30f0b', '西瓜有限公司'],
        #                                              ['149682f87c5dee5f89c3d7de109bc6a8', '北京网易传媒有限公司'],
        #                                              ['hbd7c8349ac51ac9be5d6b85cb332d26', 'mobike (Hong Kong) Limited'],
        #                                              ['haa99c369291e4a81a2390028ca798c8', '互聯網嘉（香港）有限公司'],
        #                                              ['h6aa61584047a781ebeeaff07cc5aaa5', '美團雲有限公司'],
        #                                              ['hb986cd0658f965ef355866f413e2ba9', '美團香港有限公司'],
        #                                              ['9c667e37619a31d2c216ef320a313930', '北京美团金融科技有限公司'],
        #                                              ['h4d7e2a438c2e1ba76339f8eaa339287', '袋鼠有限公司'],
        #                                              ['92da5739c12d80eeb41b988ff2c837db', '北京新美大科技有限公司'],
        #                                              ['a489c6c94f816e3a74b4b7cfa9041f44', '上海三快科技有限公司'],
        #                                              ['92da5739c12d80eeb41b988ff2c837db', '北京新美大科技有限公司'],
        #                                              ['f8a5ed873cd4ff56a34fbba6503530f7', '深圳三快商业保理有限公司'],
        #                                              ['h5479973bda97a48d62fffeefc922430', 'Myriad Dragon HK Limited'],
        #                                              ['1985ea61246a633310a047bd9e93d7d8', '点亮生活（天津）资产管理合伙企业（有限合伙）'],
        #                                              ['52954a561b419a6c21c4d8238b263519', '点亮（天津）资产管理有限公司'],
        #                                              ['1dac5c8be14a5cf94a4c6ca3ae4876d3', '上海汉涛信息咨询有限公司'],
        #                                              ['74be72b487cfef99738c9d863b78538a', '上海路团科技有限公司'],
        #                                              ['c1eff2bf8e9ec24eb0b97f1442615e57', '天津安特厨科技有限公司'],
        #                                              ['eb1820e9f126d43c8b63ddc2f8bf84aa', '宁波梅山保税港区美兴投资管理有限公司'],
        #                                              ]
        # self.excel_path = 'file/meituan.xlsx'

        self.invested_enterprise_information_list = [['h1470833b992ecf08d27e01e1ebf8d51','香港網易互動娛樂有限公司'],
                                                     ['c17e212e68f6464d0813d5def14fa1a2','广州网易计算机系统有限公司'],
                                                     ['g7ec049883a1eba9d5a1e862fb561fa2','HQG, INC.'],
                                                     ['h401da450e33860f1fd7f77a6bc1c711','網易（香港）有限公司'],
                                                     ['h25f8298bba1b7cd51307fb9129e3134','雲村有限公司'],
                                                     ['0eed34db06bc656834163abe671b08ff','北京网易有道计算机系统有限公司'],
                                                     ['h8dc0c8cd4ccfb0a43e4c91cc5d3dbeb','網易傳媒（香港）有限公司'],
                                                     ['149682f87c5dee5f89c3d7de109bc6a8','北京网易传媒有限公司'],
                                                     ['h09cfcc23c8a42fc914417682552967b','有道（香港）有限公司'],
                                                     ['0eed34db06bc656834163abe671b08ff','北京网易有道计算机系统有限公司'],]
        self.excel_path = 'file/net_ease.xlsx'



    def start(self, create_flag):
        if create_flag:
            self.excel_create()
        self.login()
        sleep(3)
        for company_information_list in self.invested_enterprise_information_list:
            self.get_single_company_info(company_information_list)

    def start_backup(self):
        self.excel_create()
        self.login()
        company_list = (['', self.information_list[1]], ['', ''], ['', ''])

        self.excel_store_information('投资关系', company_list)
        self.excel_store_information('投资关系', self.get_invested_enterprise_information())
        sleep(1)
        self.excel_store_information('软件著作', company_list)
        self.excel_store_information('软件著作', self.get_software_copyright_information())
        sleep(1)
        self.excel_store_information('app', company_list)
        self.excel_store_information('app', self.get_app_information())
        sleep(1)
        self.excel_store_information('小程序', company_list)
        self.excel_store_information('小程序', self.get_wechat_mini_program_information())
        sleep(1)
        self.excel_store_information('公众号', company_list)
        self.excel_store_information('公众号', self.get_wechat_public_account())
        sleep(1)
        self.excel_store_information('域名', self.get_website())
        sleep(1)

    def get_single_company_info(self, company_information_list):

        # append data to invested_enterprise_information_list like ['company name hash','company name'] and copy to
        # information_list
        self.information_list = company_information_list
        # to insert company name in table
        name_flag_list = (['', self.information_list[1]], ['', ''], ['', ''])

        invested_enterprise_information = self.get_invested_enterprise_information()
        self.excel_store_information('投资关系', name_flag_list)
        self.excel_store_information('投资关系', invested_enterprise_information)

        #  store invested_enterprise_information into invested_enterprise_information_list
        company_name_list = invested_enterprise_information[0]
        company_name_hash_list = invested_enterprise_information[1]
        for (company_name_hash, company_name) in zip(company_name_hash_list, company_name_list):
            self.invested_enterprise_information_list.append([company_name_hash, company_name])

        sleep(1)
        self.excel_store_information('软件著作', name_flag_list)
        self.excel_store_information('软件著作', self.get_software_copyright_information())
        sleep(1)
        self.excel_store_information('app', name_flag_list)
        self.excel_store_information('app', self.get_app_information())
        sleep(1)
        self.excel_store_information('小程序', name_flag_list)
        self.excel_store_information('小程序', self.get_wechat_mini_program_information())
        sleep(1)
        self.excel_store_information('公众号', name_flag_list)
        self.excel_store_information('公众号', self.get_wechat_public_account())
        sleep(1)
        self.excel_store_information('域名', self.get_website())

    def excel_create(self):
        # if not openpyxl.load_workbook(filename=self.excel_path):
        workbook = openpyxl.Workbook()
        row = 1
        ws = workbook.create_sheet("域名")
        order = ['网站名称', '网址', '域名', '审核日期']
        for col, entry in enumerate(order, start=1):
            ws.cell(row=row, column=col, value=entry)

        ws = workbook.create_sheet("软件著作")
        order = ['软件名称', '版本号', '发布日期', '软件简称', '登记批准日期']
        for col, entry in enumerate(order, start=1):
            ws.cell(row=row, column=col, value=entry)

        ws = workbook.create_sheet("app")
        order = ['名称', '分类', '当前版本', '简介']
        for col, entry in enumerate(order, start=1):
            ws.cell(row=row, column=col, value=entry)

        ws = workbook.create_sheet("小程序")
        order = ['名称', '分类']
        for col, entry in enumerate(order, start=1):
            ws.cell(row=row, column=col, value=entry)

        ws = workbook.create_sheet("公众号")
        order = ['微信公众号', '微信id', '简介']
        for col, entry in enumerate(order, start=1):
            ws.cell(row=row, column=col, value=entry)

        ws = workbook.create_sheet("投资关系")
        order = ['被投资企业名称', 'href', '投资比例']
        for col, entry in enumerate(order, start=1):
            ws.cell(row=row, column=col, value=entry)

        workbook.save(self.excel_path)

    def excel_store_information(self, sheet_name, information):
        workbook = openpyxl.load_workbook(filename=self.excel_path)
        worksheet = workbook[sheet_name]
        row = worksheet.max_row + 1
        if isinstance(information, tuple):
            tmp = row
            col = 1
            for single_col in information:
                for row, entry in enumerate(single_col, start=row):
                    worksheet.cell(row=row, column=col, value=entry)
                    row += 1
                row = tmp
                col += 1
        elif isinstance(information, list):
            row = worksheet.max_row + 1
            for col, entry in enumerate(information, start=1):
                worksheet.cell(row=row, column=col, value=entry)
        else:
            print("is not list or tuple!")
        workbook.save(self.excel_path)

    def login(self):
        self.driver.get(
            'https://graph.qq.com/oauth2.0/show?which=Login&display=pc&response_type=code&client_id=101188807&redirect_uri=http%3A%2F%2Fwww.qichacha.com%2Fuser_callbackqq&state=4fe189a24145c0848fb69c3a1da72c51&scope=get_user_info,add_share')
        self.driver.switch_to.frame('ptlogin_iframe')
        self.driver.find_element(By.ID, "img_out_1356242944").click()
        self.driver.switch_to.window(self.driver.window_handles[0])
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "header navi-header box-shadow "))
            )
        except Exception as ex:
            print("the page load time out (10s)")

    def get_information(self, url, xpath_list, page_exit_flag):
        # the url_list will be depart by page variable
        if page_exit_flag:
            page = 1
            page_flag = True
            self.driver.get(url[0] + str(page) + url[1])
            for xpath in xpath_list:
                if self.driver.find_elements_by_xpath(xpath[0]):
                    while page_flag:
                        if self.driver.find_elements_by_xpath(xpath[0]):
                            for single_xpath in xpath:
                                yield self.driver.find_elements_by_xpath(single_xpath)
                            page = page + 1
                            self.driver.get(url[0] + str(page) + url[1])
                        else:
                            page_flag = False

    def get_information_backup(self, url, xpath_list, page_exit_flag):
        # the url_list will be depart by page variable
        if page_exit_flag:
            page = 1
            page_flag = True
            while page_flag == True:
                self.driver.get(url[0] + str(page) + url[1])
                if self.driver.find_elements_by_xpath(xpath_list[0]):
                    for xpath in xpath_list:
                        yield self.driver.find_elements_by_xpath(xpath)
                    page = page + 1
                else:
                    page_flag = False
        else:
            self.driver.get(url)
            for xpath in xpath_list:
                yield self.driver.find_elements_by_xpath(xpath)

    def get_invested_enterprise_information(self):

        xpath_list1 = ["//td[@width='235']//span[@class='seo font-14']",
                       "//td[@width='235']//a[@target='_blank']", "//td[@width='82']"]

        xpath_list2 = ["//td[@width='']//span[@class='seo font-14']",
                       "//td[@width='']//a[@target='_blank']", "//td[@width='110']"]

        xpath_list = [xpath_list1, xpath_list2]

        url = ["https://www.qcc.com/company_getinfos?unique=" + self.information_list[0] + "&companyname=" +
               self.information_list[1] + "&p=", "&tab=base&box=touzi"]

        name_of_invested_enterprise_list = []
        invested_enterprise_href_list = []
        investment_ratio_list = []

        # name_of_invested_enterprise, invested_enterprise_href, investment_ratio = self.get_information(url, xpath_list,
        # True)
        information = self.get_information(url, xpath_list, True)
        count = 1
        xpath_list_length = len(xpath_list[0])
        for tem_next in information:
            if count % xpath_list_length == 1:
                for i in tem_next:
                    name_of_invested_enterprise_list.append(i.text)
            elif count % xpath_list_length == 2:
                for i in tem_next:
                    invested_enterprise_href_list.append(i.get_attribute('href'))
            elif count % xpath_list_length == 0:
                for i in tem_next:
                    investment_ratio_list.append(i.text)
            count += 1

        # get the text from element
        return self.get_control_company_information(name_of_invested_enterprise_list,
                                                    invested_enterprise_href_list,
                                                    investment_ratio_list)

    # delete the company whose percentage less than 50%
    def get_control_company_information(self, name_of_invested_enterprise_list,
                                        invested_enterprise_href_list,
                                        investment_ratio_list):
        if len(investment_ratio_list) == len(name_of_invested_enterprise_list):
            index = len(investment_ratio_list)
        else:
            raise RuntimeError('two list length not equal')
        i = 0
        while i < index:
            percentage = investment_ratio_list[i]
            if "%" in percentage:
                float_percentage = float(percentage.strip("%")) / 100
                if float_percentage < 0.5:
                    name_of_invested_enterprise_list.pop(i)
                    invested_enterprise_href_list.pop(i)
                    investment_ratio_list.pop(i)
                    index = index - 1
                    continue
            else:
                # raise RuntimeError('the object is not a percentage object!')
                print(name_of_invested_enterprise_list[i]+'the object is not a percentage object!')

            i = i + 1

        # get the unique
        index = len(invested_enterprise_href_list)
        i = 0
        while i < index:
            invested_enterprise_href_list[i] = re.findall(r"firm/(.+?).html", invested_enterprise_href_list[i])[0]
            # print(invested_enterprise_href_list[i])
            i = i + 1

        return name_of_invested_enterprise_list, invested_enterprise_href_list, investment_ratio_list

    def get_software_copyright_information(self):

        xpath_list1 = ["//td[@width='278']", "//td[@width='9%']", "//td[@width='140']", "//td[@width='16%']",
                       "//td[@width='150']"]
        xpath_list = [xpath_list1]
        url = ["https://www.qcc.com/company_getinfos?unique=" + self.information_list[0] + "&companyname=" +
               self.information_list[1] + "&p=", "&tab=assets&box=rjzzq"]

        # url = ["file:///C:/Users/yangzhanan/Desktop/html/softwareCopyright/" + self.information_list[0] + "&companyname=" +
        #        self.information_list[1] + "&p=", "&tab=assets&box=rjzzq.html"]
        software_name_list = []
        version_number_list = []
        pubilc_time_list = []
        software_abbreviation_list = []
        register_time = []

        information = self.get_information(url, xpath_list, True)

        count = 1
        xpath_list_length = len(xpath_list[0])
        for tem_next in information:
            if count % xpath_list_length == 1:
                for i in tem_next:
                    software_name_list.append(i.text)
            elif count % xpath_list_length == 2:
                for i in tem_next:
                    version_number_list.append(i.text)
            elif count % xpath_list_length == 3:
                for i in tem_next:
                    pubilc_time_list.append(i.text)
            elif count % xpath_list_length == 4:
                for i in tem_next:
                    software_abbreviation_list.append(i.text)
            elif count % xpath_list_length == 0:
                for i in tem_next:
                    register_time.append(i.text)
            count += 1

        return software_name_list, version_number_list, pubilc_time_list, software_abbreviation_list, register_time

    def get_app_information(self):
        xpath_list1 = ["//img[@style='width: 66px;']//following::td[1]",
                       "//img[@style='width: 66px;']//following::td[2]",
                       "//img[@style='width: 66px;']//following::td[3]",
                       "//img[@style='width: 66px;']//following::span[1]"]
        xpath_list = [xpath_list1]

        url = ["https://www.qcc.com/company_getinfos?unique=" + self.information_list[0] + "&companyname=" +
               self.information_list[1] + "&p=", "&tab=assets&box=app"]
        app_name_list = []
        classification_list = []
        current_version_list = []
        brief_introduction_list = []

        information = self.get_information(url, xpath_list, True)
        count = 1
        xpath_list_length = len(xpath_list[0])
        for tem_next in information:
            if count % xpath_list_length == 1:
                for i in tem_next:
                    app_name_list.append(i.text)
            elif count % xpath_list_length == 2:
                for i in tem_next:
                    classification_list.append(i.text)
            elif count % xpath_list_length == 3:
                for i in tem_next:
                    current_version_list.append(i.text)
            elif count % xpath_list_length == 0:
                for i in tem_next:
                    brief_introduction_list.append(i.text)
            count += 1

        return app_name_list, classification_list, current_version_list, brief_introduction_list

    def get_wechat_mini_program_information(self):
        xpath_list1 = ["//img[@style='width: 66px;']//following::td[1]",
                       "//img[@style='width: 66px;']//following::td[2]"]
        xpath_list = [xpath_list1]
        url = ["https://www.qcc.com/company_getinfos?unique=" + self.information_list[0] + "&companyname=" +
               self.information_list[1] + "&p=", "&tab=assets&box=wp"]
        mini_program_nam_list = []
        classification_list = []

        information = self.get_information(url, xpath_list, True)

        count = 1
        xpath_list_length = len(xpath_list[0])
        for tem_next in information:
            if count % xpath_list_length == 1:
                for i in tem_next:
                    mini_program_nam_list.append(i.text)
            elif count % xpath_list_length == 0:
                for i in tem_next:
                    classification_list.append(i.text)
            count += 1

        return mini_program_nam_list, classification_list

    def get_wechat_public_account(self):
        xpath_list1 = ["//img[@style='width: 66px;']//following::td[1]",
                       "//img[@style='width: 66px;']//following::td[2]",
                       "//img[@style='width: 66px;']//following::td[4]"]
        xpath_list = [xpath_list1]
        url = ["https://www.qcc.com/company_getinfos?unique=" + self.information_list[0] + "&companyname=" +
               self.information_list[1] + "&p=", "&tab=assets&box=wechat"]

        name_list = []
        id_list = []
        brief_introduction_list = []
        information = self.get_information(url, xpath_list, True)
        count = 1

        xpath_list_length = len(xpath_list[0])

        for tem_next in information:
            if count % xpath_list_length == 1:
                for i in tem_next:
                    name_list.append(i.text)
            elif count % xpath_list_length == 2:
                for i in tem_next:
                    id_list.append(i.text)
            elif count % xpath_list_length == 0:
                for i in tem_next:
                    brief_introduction_list.append(i.text)
            count += 1

        return name_list, id_list, brief_introduction_list

    def get_website(self):

        xpath_list1 = ["//td[@width='15%']", "//td[@width='18%']//a", "//td[@width='13%']", "//td[@width='103']"]
        xpath_list = [xpath_list1]
        website_name_list = []
        website_url_list = []
        website_domain_list = []
        audit_date_list = []
        # record official website
        official_website = self.get_official_website()
        if not official_website:
            website_name_list.append(self.information_list[1] + ":官网暂无")
            website_url_list.append("null")
            website_domain_list.append("null")
            audit_date_list.append("null")
        else:
            website_name_list.append(self.information_list[1] + "官网")
            website_url_list.append(re.findall(r'(https|http)://(.+)/', official_website)[0][1])
            website_domain_list.append(re.findall(r'(https|http)://www.(.+)/', official_website)[0][1])
            audit_date_list.append("unknown")

        url = ["https://www.qcc.com/company_getinfos?unique=" + self.information_list[0] + "&companyname=" +
               self.information_list[1] + "&p=", "&tab=assets&box=website"]

        information = self.get_information(url, xpath_list, True)
        count = 1
        xpath_list_length = len(xpath_list[0])
        for tem_next in information:
            if count % xpath_list_length == 1:
                for i in tem_next:
                    website_name_list.append(i.text)
            elif count % xpath_list_length == 2:
                for i in tem_next:
                    website_url_list.append(i.text)
            elif count % xpath_list_length == 3:
                for i in tem_next:
                    website_domain_list.append(i.text)
            elif count % xpath_list_length == 0:
                for i in tem_next:
                    audit_date_list.append(i.text)
            count += 1

        for index, domain in enumerate(website_domain_list):
            if index != 0:
                if domain == website_domain_list[0]:
                    website_name_list[0] = website_name_list[0] + "/" + website_name_list[
                        website_domain_list.index(domain)]
                    website_name_list.pop(index)
                    website_url_list.pop(index)
                    website_domain_list.pop(index)
                    audit_date_list.pop(index)

        return website_name_list, website_url_list, website_domain_list, audit_date_list

    def get_official_website(self):
        url = "https://www.qcc.com/firm/" + self.information_list[0] + ".html"
        self.driver.get(url)
        self.driver.switch_to.window(self.driver.window_handles[0])
        # return self.driver.find_element_by_xpath("a[@data-original-title='进入官网']").get_attribute('href')
        try:
            return WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='company-top']/div[2]/div[2]/div[5]/div[2]/span[3]/a"))
            ).get_attribute('href')
        except:
            return False


if __name__ == "__main__":
    t1 = information_gather()
    t1.start(create_flag=True)
